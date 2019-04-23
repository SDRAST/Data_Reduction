# -*- coding: utf-8 -*-
"""
This provides functions for making and reading spreadsheets from GAVRT t-files.

These may now be obsolete because we don't turn data in the DSS-28 EAC database
into spreadsheet.

This assumes that only one IF-mode channel is used, since solar emission is so
strong.

Functions::
  column_ID_dict(worksheet)
    Create a dictionary of column number indexed by column name
  create_boresight_analysis_sheet(wb)
    Create a sheet for processing calibration data
  create_map_analysis_sheet(wb)
    Create a sheet for processing map data
  fill_map_sheet(wb,sheetname,dxdecs,ddecs,data)
    Fill a spreadsheet page with map data
  get_timed_data_sheet(wb,fkeys,pkeys,date_nums,ras,decs,azs,elevs)
    See if the timed data sheet exist and create it if not
  get_column(ws, column_name)
    Get the data as a list from a column identified by its name.
  get_map_sheet(wb,sheetname)
    Get map data from a map sheet
  get_rows(sh,colname,value)
    Get a list of rows (each a list of cells) with a given value in a column
  get_timed_data(wb,sheetname)
    Get the data from Tsys vs time data sheet
  load_data_sheet(name)
    load metadata into the observation session "Metadata" worksheet
"""
from matplotlib.pylab import date2num, num2date
from numpy import array, zeros
from os.path import exists
from pylab import *

from openpyxl import load_workbook, Workbook
from openpyxl.reader.excel import InvalidFileException
from support.excel import *
from Data_Reduction import nearest_index
#from support.excel import set_column_dimensions
from support.lists import unique

diag = True

def get_timed_data_sheet(wb,fkeys,pkeys,date_nums,ras,decs,azs,elevs):
  """
  See if the timed data sheet exist and create it if not

  A timed data sheet is basically the power VtoF data as a function
  of time (rows) and frequency/polarization (columns).  This initiates
  the sheet but does not fill it.  fill_map_sheet() does that.

  @param wb : the spreadsheet
  @type  wb : openpyxl.Wotkbook() instance

  @param fkeys : frequency-based keys into the data
  @type  fkeys : list of float

  @param pkeys : polarization-based keys
  @type  pkeys : list of str

  @param date_nums : datetime for each sample
  @type  date_nums : matplotlib date numbers

  @param ras : right ascensions
  @type  ras : list of floats (hr)

  @param decs : declinations
  @type  decs : list of floats (deg)

  @param azs : azimuths
  @type  azs : list of floats
  
  @param elevs : elevations
  @type  elevs : list of float

  @return: datasheet
  """
  timedata_sheet = wb.get_sheet_by_name("TimeData")
  if timedata_sheet == None:
    # Doesn't exist.  Make it.
    if diag:
      print "Trying to create a new data sheet"
    try:
      timedata_sheet = wb.create_sheet(1)
    except Exception,details:
      print "Could not create sheet"
      print details
      return None
    timedata_sheet.title = "TimeData"
    if diag:
      print timedata_sheet,"has been given the name",\
            timedata_sheet.title
    # make column titles
    column_names = ["Date/Time","R.A.","Decl.","Az","Elev"]
    for freq in fkeys:
      for pol in pkeys:
        column_names.append( ("%5.2f-%3s" % (freq,pol)).strip() )
    for index in range(len(column_names)):
      timedata_sheet.cell(row = 0, column = index).value = column_names[index]
    # Fill the columns
    for count in range(len(date_nums)):
      timedata_sheet.cell(row = count+1,
               column = get_column_id(timedata_sheet,"Date/Time")).value = \
                                              num2date(date_nums[count])
      timedata_sheet.cell(row = count+1,
               column = get_column_id(timedata_sheet,"R.A.")).value = \
                                              ras[count]
      timedata_sheet.cell(row = count+1,
               column = get_column_id(timedata_sheet,"Decl.")).value = \
                                              decs[count]
      timedata_sheet.cell(row = count+1,
               column = get_column_id(timedata_sheet,"Elev")).value = \
                                              elevs[count]
      timedata_sheet.cell(row = count+1,
               column = get_column_id(timedata_sheet,"Az")).value = \
                                              azs[count]
  if diag:
      print timedata_sheet.title,"sheet exists"
  return timedata_sheet

def fill_map_sheet(wb,sheetname,dxdecs,ddecs,data):
  """
  Fill a spreadsheet page with map data

  The spreadsheet has one page per frequency/polarization,
  as a function of cross-dec offset (columns) and dec offset (rows)

  @param wb : spreadsheet
  @type  wb : openpyxl.Workbook() instance

  @param sheetname : name based on frequency and polarization
  @type  sheetname : str

  @param dxdecs : cross-declination offsets
  @type  dxdecs : list of floats

  @param ddecs : declination offsets
  @type  ddecs : list of floats

  @param data : 2D image
  @type  data : numpy array of float

  @return: openpyxl.Worksheet() instance
  """
  mapsheet = wb.get_sheet_by_name(sheetname)
  if mapsheet == None:
    # Doesn't exist, make it.
    if diag:
      print "Attempting to create worksheet",sheetname
    try:
      mapsheet = wb.create_sheet(-1)
    except Exception, details:
      print "Could not create sheet"
      print details
      return None
    if diag:
      print "Created worksheet",mapsheet,"named",sheetname
    mapsheet.title = sheetname
  else:
    if diag:
      print "Opened sheet",mapsheet.title
  # fill the top row with cross-declination offsets
  for count in range(len(dxdecs)):
    mapsheet.cell(row = 0, column = count+1).value = dxdecs[count]
  # fill the left column with declination offsets
  for count in range(len(ddecs)):
    mapsheet.cell(row = count+1, column = 0).value = ddecs[\linebreakcount]
  # fill in the map data
  for dxdec_index in range(len(dxdecs)):
    for ddec_index in range(len(ddecs)):
      mapsheet.cell(column = dxdec_index+1, row = ddec_index+1).value = \
                  float(data[dxdec_index,ddec_index])
  return mapsheet

def get_map_sheet(wb,sheetname):
  """
  Get map data from a map sheet
  """
  mapsheet = wb.get_sheet_by_name(sheetname)
  if mapsheet == None:
    return None
  else:
    dxdecs = []
    for column in range(1,mapsheet.get_highest_column()):
      dxdecs.append(mapsheet.cell(row=0, column=column).value)Wedne
    ddecs = []
    for row in range(1,mapsheet.get_highest_row()):
      ddecs.append(mapsheet.cell(row=row, column=0).value)
    data = zeros((len(dxdecs),len(ddecs)))
    for column in range(1,mapsheet.get_highest_column()):
      for row in range(1,mapsheet.get_highest_row()):
        data[row-1,column-1] = mapsheet.cell(row=row,column=column).value
  return array(dxdecs),array(ddecs),data
  
def load_data_sheet(name):
  """
  load metadata into the observation session "Metadata" worksheet

  @param name : workbook name
  @type  name : str

  @return: (Workbook() instance, Worksheet() instance, column names and a
           dictionary of column numbers indexed by name`
  """
  try:
    wb = load_workbook(name)
  except IOError, details:
    print "Loading spreadsheet failed with IO error."
    print details
    sys.exit(1)
  except AttributeError, details:
    print "Loading spreadsheet failed with attribute error"
    print details
    sys.exit(1)
  except InvalidFileException, details:
    print "File does not exist.  Creating it."
    wb = Workbook()
    obs_ws = wb.get_active_sheet()
    obs_ws = create_metadata_sheet(obs_ws,"Metadata")
  else:
    sheet_names = wb.get_sheet_names()
    print "Sheets:",sheet_names
    obs_ws = wb.get_sheet_by_name('Metadata')
  # make a reverse lookup table
  obs_col_names, meta_column = column_ID_dict(obs_ws)\linebreak
  return wb, obs_ws, obs_col_names, meta_column
  
def get_timed_data(wb,sheetname):
  """
  Get the data from Tsys vs time data sheet
  """
  sheet = wb.get_sheet_by_name(sheetname)
  date_nums = date2num(array(get_column(sheet,"Date/Time")))
  ras       = get_column(sheet,"R.A.")
  decs      = get_column(sheet,"Decl.")
  azs       = get_column(sheet,"Az")
  elevs     = get_column(sheet,"Elev")
  col_name_dict = get_column_names(sheet)
  colnames = col_name_dict.values()
  colnames.remove("Date/Time")
  colnames.remove("R.A.")
  colnames.remove("Decl.")
  colnames.remove("Az")
  colnames.remove("Elev")
  tsys = {}
  for colname in colnames:
    tsys[colname] = get_column(sheet,colname)
  return date_nums, ras, decs, azs, elevs, tsys

def create_boresight_analysis_sheet(wb):
  """
  Create a sheet for processing calibration data
  """
  analsh = wb.create_sheet(index=-1,title="Analysis")
  analsh.cell(column= 0,row=0).value = "Source"
  analsh.cell(column= 1,row=0).value = "Freq"
  analsh.cell(column= 2,row=0).value = "Pol"
  analsh.cell(column= 3,row=0).value = "IF mode"Wedne
  analsh.cell(column= 4,row=0).value = "Dec Start"
  analsh.cell(column= 5,row=0).value = "Dec End"
  analsh.cell(column= 6,row=0).value = "XDec Start"
  analsh.cell(column= 7,row=0).value = "XDec End"
  analsh.cell(column= 8,row=0).value = "Off Start"
  analsh.cell(column= 9,row=0).value = "Off End"
  analsh.cell(column=10,row=0).value = "On Start"
  analsh.cell(column=11,row=0).value = "On End"
  analsh.cell(column=12,row=0).value = "Flux"
  analsh.cell(column=13,row=0).value = "Flux Ref"
  analsh.cell(column=14,row=0).value = "Counts"
  analsh.cell(column=15,row=0).value = "Gain"
  analsh.cell(column=16,row=0).value = "Elev"
  return analsh

def create_map_analysis_sheet(wb):
  """
  Create a sheet for processing map data
  """
  analsh = wb.create_sheet(index=-1,title="Analysis")
  analsh.cell(column= 0,row=0).value = "Source"
  analsh.cell(column= 1,row=0).value = "Freq"
  analsh.cell(column= 2,row=0).value = "Pol"
  analsh.cell(column= 3,row=0).value = "IF mode"
  analsh.cell(column= 4,row=0).value = "Dec Offset"
  analsh.cell(column= 5,row=0).value = "XDec Offset"
  analsh.cell(column= 6,row=0).value = "Dec Height"
  analsh.cell(column= 7,row=0).value = "XDec Width"
  analsh.cell(column= 8,row=0).value = "Dec Step"
  analsh.cell(column= 9,row=0).value = "XDec Step"
  analsh.cell(column=10,row=0).value = "Integration"
  analsh.cell(column=11,row=0).value = "Gain"
  analsh.cell(column=12,row=0).value = "Elev"
  return analsh

def create_metadata_sheet(obs_ws,title):
  obs_ws.title = title
  obs_ws.cell(column=0,row=0).value = "File"
  obs_ws.cell(column=1,row=0).value = "Freq"
  obs_ws.cell(column=2,row=0).value = "Pol"
  obs_ws.cell(column=3,row=0).value = "IF mode"
  obs_ws.cell(column=4,row=0).value = "First"
  obs_ws.cell(column=5,row=0).value = "Last"
  obs_ws.cell(column=6,row=0).value = "Start"
  obs_ws.cell(column=7,row=0).value = "Stop"
  obs_ws.cell(column=8,row=0).value = "Attn"
  return obs_ws
  
def create_boresight_fit_sheet(wb,name):
  """
  """
  sheet = wb.create_sheet(index=-1,title=name)
  sheet.cell(column= 0,row=0).value = "Freq"
  sheet.cell(column= 1,row=0).value = "Pol"
  sheet.cell(column= 2,row=0).value = "IF mode"
  sheet.cell(column= 3,row=0).value = "Dec First"
  sheet.cell(column= 4,row=0).value = "Dec Last"
  sheet.cell(column= 5,row=0).value = "XDec First"
  sheet.cell(column= 6,row=0).value = "XDec Last"
  return sheet

def get_data_segments(analsh, col_ID,
                      time_array, data_lists,
                      start_time_column, end_time_column,
                      row):
  """
  Extract data segments from lists of timed data

  The start and stop times are taken from 'row' in 'analsh'.  'time_array',
  'elevs' and 'data' all have the same length.  The start and stop times
  are in the columns given by name.

  Note
  ====
  polyfit cannot handle datetime.datetime() objects

  @param analsh : boresight "Analysis" worksheet
  @type  analsh : openpyxl.Worksheet() instance

  @param col_ID : column numbers indexed by column name
  @type  col_ID : dict

  @param time_array : time data
  @type  time_array : numpy array of matplotlib time numbers

  @param data_lists : tuple of lists of data to be extracted
  @type  data_lists : tuple of lists of float

  @param start_time_column : name of column with start times
  @type  start_time_column : str

  @param end_time_column : name of column with end times
  @type  end_time_column : str

  @param row : row number from which to take start and stop times
  @type  row : int

  @return: time_array, list of data_array, index of first row, index of last
  """
  start_time = analsh.cell(row=row, column=col_ID[start_time_column]).value
  end_time   = analsh.cell(row=row, column=col_ID[end_time_column]).value
  start_index = nearest_index(time_array,start_time)
  end_index   = nearest_index(time_array,end_time)
  times = date2num(time_array[start_index:end_index])
  data_arrays = []
  for data in data_lists:
    data_array = array( data[start_index:end_index])
    data_arrays.append(data_array)
  return times, data_arrays, start_index, end_index

def get_segment_filename(project_path,date_info,sheet):
  """
  Gets the spreadsheet filename for a data segment

  @param date_info : result from DatesTimes.get_date()
  @type  date_info : list (str, int, int, int, int)

  @param sheet : data segment metasheet name
  @type  sheet : str

  @return: relative_path (str)
  """
  session_path = project_path+"Data/"+date_info[0]
  datestr = "%4d-%03d" % (date_info[1],date_info[4])
  name_parts = sheet.split('-')
  name_parts.insert(1,datestr)
  spreadsheet = '-'.join(name_parts)+'.xlsx'
  segment_filename = session_path+"/" + spreadsheet
  return segment_filename, name_parts

def create_gridded_maps(wb,map_meta_sh,filename_dict):
  """
  Create gridded maps for each receiver channel

  Notes
  =====
  "Map-HHMM-HHMM" sheets have meta data for the observations of this segment
  "Analysis" sheets have meta data for the maps obtained from calibrations

  @param wb : map spreadsheet
  @type  wb : openpyxl.Workbook() instance

  @param map_meta_sh : worksheet with map metadata
  @type  map_meta_sh : openpyxl.worksheet.Worksheet() instance

  @param filename_dict : t-file name indexed by frequency and polarization
  @type  filename_dict : dict

  @return: None
  """
  # get the time from the meta data sheet
  first_time = map_meta_sh.cell(row=1,
                                column=OP.get_column_id(map_meta_sh,
                                                     'Start')).value
  last_time  = map_meta_sh.cell(row=1,
                                column=OP.get_column_id(map_meta_sh,
                                                     'Stop')).value
  # get map parameters
  map_anal_sh = wb.get_sheet_by_name("Analysis")
  #data_sheet = wb.get_sheet_by_name("TimeData")
  date_nums, ras, decs, azs, elevs, tsys = get_timed_data(wb,"TimeData")
  fkeys = unique(get_column(map_anal_sh,'Freq'))
  pol_col = OP.get_column_id(map_anal_sh,'Pol')
  for f in fkeys:
    # This gets all the rows which have this frequency
    rows = get_rows(map_anal_sh,'Freq',f)
    if rows:
      for row in rows:
        # This steps through the polarizations
        p = row[pol_col].value
        # set the grid parameters
        xdec_off   = row[OP.get_column_id(map_anal_sh,'XDec Offset')].value
        dec_off    = row[OP.get_column_id(map_anal_sh,'Dec Offset')].value
        map_height = row[OP.get_column_id(map_anal_sh,'Dec Height')].value
        map_width  = row[OP.get_column_id(map_anal_sh,'XDec Width')].value
        xdec_step  = row[OP.get_column_id(map_anal_sh,'XDec Step')].value
        dec_step   = row[OP.get_column_id(map_anal_sh,'Dec Step')].value
        xi = arange(xdec_off - map_width/2,
                    xdec_off + map_width/2  + xdec_step, xdec_step/2)
        yi = arange(dec_off  - map_height/2,
                    dec_off  + map_height/2 + dec_step,  dec_step/2)
        # get the data from the timed data page
        column_name = pagename(f,p)
        if diag:
          print "Processing counts column",column_name
        dxdecs,ddecs = center_data(date_nums, ras, decs,
                                   "sun", dss28)
        zi = griddata(dxdecs,ddecs,tsys[column_name],xi,yi)
        mapsheet = fill_map_sheet(wb,pagename(f,p),xi,yi,zi)
        set_column_dimensions(mapsheet)
  return

def plot_track(map_type, project_path, date_info,
               wb, sheet, first_time, last_time,
               fkeys, filename_dict,
               fig,
               dt_start, dt_stop):
  """
  Plot the positions of the map samples

  @param map_type : 'azel', 'radec', 'xdecdec'
  @type  map_type " str

  @param date_info : response from get_date(): ('2012-05-20', 2012, 5, 20, 141)
  @type  date_info : (str, int, int, int, int)

  @param wb : spreadsheet
  @type  wb : openpyxl.Workbook() instance

  @param sheet : currently selected data segment; could be None
  @type  sheet : str

  @param fkeys : frequencies (GHz) used in this session
  @type  fkeys : list of float

  @param filename_dict : index into t-files by freq and pol
  @type  filename_dict : dict[][]

  @param fig : current figure number
  @type  fig : int

  @param dt_start : start of the observing session
  @type  dt_start : datetime.datetime() instance

  @param dt_stop : end of the observing session
  @type  dt_stop : datetime.datetime() instance
  """
  if sheet:
    print "Processing segment sheet",sheet
    seg_file, name_parts = get_segment_filename(project_path,date_info,sheet)
    if exists(seg_file):
      if diag:
        print "Loading",seg_file
      wb = load_workbook(seg_file)
      date_nums, ras, decs, azs, elevs, tsys_dict = get_timed_data(wb,
                                                                  "TimeData")
    else:
      if diag:
        print seg_file,"does not exist"
      seg_file = None
      return fig
  else:
    if diag:
      print "Processing the entire session"
  # plot the elevation vs azimuth, one figure per frequency
  for f in fkeys:
    fig += 1
    figure(fig)
    subpl = 0
    for p in filename_dict[f].keys():
      subpl += 1
      subplot(1,2,subpl)
      if seg_file:
        tsys = tsys_dict[pagename(f,p)]
      else:
        if sheet:
          # ... but no data file
          start = first_time
          stop  = last_time
        else:
          # all the data
          start = dt_start[p]
          stop  = dt_stop[p]
          filename = filename_dict[f][p]
          filepath = "../"+date_info[0]+"/" + filename
          data, labels, date_nums, ras, decs, azs, elevs, tsys = \
                load_tfile_data(filepath,start=start,stop=stop)
      meantime = num2date((date_nums[0]+date_nums[-1])/2.)
      doy = day_of_year(mean_time.year, mean_time.month, mean_time.day)
      title_str = "%4d/%03d %5.2f GHz %3s" % (mean_time.year,doy, f, p)
      if map_type == 'azel':
        plot_azel(azs,elevs,title_str)
        suffix = "-azel.png"
      elif map_type == 'radec':
        plot_ra_dec(ras,decs,title_str)
        suffix = "-radec.png"
      elif map_type == 'xdecdec':
        dxdecs,ddecs = center_data(date_nums,ras,decs,"sun",dss28)
        plot_xdec_dec(dxdecs,ddecs,title_str)
        suffix = "-decxdec.png"
      else:
        print "Invalid map type:",map_type
        suffix = None
    if suffix:
      savefig(("../%s/%4d-%03d_%05.2f" % (date_info[0],
                                          mean_time.year,doy,
                                          f))+suffix,
              orientation="landscape",format="png")
  return fig

def show_contoured_image(wb, meta_sheet_name, mean_time, fig):
  """
  Show false-colored contoured images for each receiver channel

  @param wb : map workbook
  @type  wb : openpyxl.Workbook() instance

  @param meta_sheet_name : metadata sheet
  @type  meta_sheet_name : str

  @param mean_time : a date/time for the map
  @type  mean_time : matplotlib datenum

  @param fig : figure number
  @type  fig : int

  @return: incremented figure number
  """
  map_meta_sh = wb.get_sheet_by_name(meta_sheet_name)
  map_anal_sh = wb.get_sheet_by_name("Analysis")
  fkeys = unique(get_column(map_meta_sh,'Freq'))
  pkeys = unique(get_column(map_meta_sh,'Pol'))
  pol_column        = OP.get_column_id(map_meta_sh,'Pol')
  start_column      = OP.get_column_id(map_meta_sh,'Start')
  stop_column       = OP.get_column_id(map_meta_sh,'Stop')
  xdec_off_column   = OP.get_column_id(map_anal_sh,'XDec Offset')
  dec_off_column    = OP.get_column_id(map_anal_sh,'Dec Offset')
  map_height_column = OP.get_column_id(map_anal_sh,'Dec Height')
  map_width_column  = OP.get_column_id(map_anal_sh,'XDec Width')
  gain_column       = OP.get_column_id(map_anal_sh,'Gain')
  for f in fkeys:
    # This steps through frequencies
    frows = get_row_numbers(map_meta_sh,'Freq',f)
    fig += 1
    thisfig = figure(fig, figsize=(8.0, 8.0))
    subpl = 0
    aper_effic = zenith_gain(f*1000.)
    if diag:
      print "Zenith gain =",aper_effic,"at",f,"GHz"
    for p in pkeys:
      # This steps through the polarizations
      row = OP.get_row_number(map_meta_sh,pol_column,p)
      first_time = map_meta_sh.cell(row=row, column=start_column).value
      last_time  = map_meta_sh.cell(row=row, column=stop_column).value
      xdec_off   = map_anal_sh.cell(row=row, column=xdec_off_column).value
      dec_off    = map_anal_sh.cell(row=row, column=dec_off_column).value
      map_height = map_anal_sh.cell(row=row, column=map_height_column).value
      map_width  = map_anal_sh.cell(row=row, column=map_width_column).value
      gain       = map_anal_sh.cell(row=row, column=gain_column).value
      mean_time  = num2date((date2num(first_time) + date2num(last_time))/2.)
      xi, yi, zi = get_map_sheet(wb,pagename(f,p))
      #     contour the gridded data
      subpl += 1
      ax = thisfig.add_subplot(1,2,subpl)
      # 1.2 converts from aperture efficiency to coupling efficiency to
      # a source which covers several to many sidelobes (depending on
      # frequency).
      Jy_to_K = 1.2*antenna_gain(aper_effic,pi*(34/2)**2)
      if f == 3.1 and p == 'RCP':
        V = arange(0.,200000.,10000.)
      elif f == 6 and p == 'RCP':
        V = arange(0.,100000., 5000.)
      elif f == 8.45 and p == 'RCP':
        V = arange(0., 50000., 2500.)
      elif f == 14 and p == 'RCP':
        V = arange(0.,  4500.,  250.)
      else:
        pass
      CS = contour( xi, yi, Jy_to_K*gain*zi, V, linewidths=0.5, colors='k')
      CS = contourf(xi, yi, Jy_to_K*gain*zi, V, cmap=plt.cm.jet)
      print "%5.2f %s %7.0f" % (f,p,(Jy_to_K*gain*zi).max())
      ax.set_aspect('equal')
      colorbar(shrink=0.6) # draw colorbar
      grid()
      xlabel("Cross-declination offset")
      if subpl == 1:
        ylabel("Declination offset")
      title("%3s" % p)
      doy = day_of_year(mean_time.year, mean_time.month, mean_time.day)
      suptitle("%4d/%03d %5.2f GHz" % (mean_time.year,doy, f))
      #     show sun-centered coords
      map_params = xdec_off, dec_off, map_width, map_height
      show_body_orientation("Sun",mean_time,map_params,solar_data)
      xlim(xdec_off-map_width/2., xdec_off+map_width/2.)
      ylim( dec_off-map_height/2., dec_off+map_height/2.)
      savefig((project_path+"Data/%s/%4d-%03d_%05.2f" % (date_info[0],
                                          mean_time.year,
                                          doy,
                                          f))+"-cntr.png",
              orientation="landscape", format="png")
      savefig((project_path+"Data/%s/%4d-%03d_%05.2f" % (date_info[0],
                                          mean_time.year,
                                          doy,
                                          f))+"-cntr.eps",
              orientation="landscape", format="eps")
  return fig

def pagename(f,p):
  """
  Create freq/pol string for use in workbook and worksheet names

  @param f : frequency
  @type  f : float

  @param p : polarization in three characters
  @type  p : str

  @return: str
  """
  return ("%5.2f-%3s" % (f,p)).strip()

def get_session_workbook(project_path, date_info):
  """
  """
  session_path = project_path+"Observations/dss28/%4d/%03d/" % (date_info[1],date_info[4])
  wb_name = "%4d-%03d.xlsx" % (date_info[1],date_info[4])
  wb_file = session_path+"/"+wb_name
  return wb_name, wb_file
