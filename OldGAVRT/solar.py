# -*- coding: utf-8 -*-
import sys
import os.path
import datetime
import logging
import openpyxl
import openpyxl.reader.excel as excel

import DatesTimes as DT
import Astronomy.solar as solar
import support.excel

logger = logging.getLogger(__name__)

def load_spreadsheet(filename, sheet='Metadata'):
  """
  Load the observations spreadsheet

  @param filename : filename of the spreadsheet

  @return: workbook instance
  """
  try:
    wb = openpyxl.load_workbook(filename)
  except IOError as details:
    logger.error("load_spreadsheet: Loading spreadsheet failed with IO error.",exc_info=True)
    return None,None
  except AttributeError as details:
    logger.error("load_spreadsheet: Loading spreadsheet failed with attribute error",
                   exc_info=True)
    return None,None
  except excel.InvalidFileException as details:
    logger.error("load_spreadsheet: File "+filename+" does not exist.",exc_info=True)
    return None,None
  else:
    sheet_names = wb.get_sheet_names()
    logger.debug("load_spreadsheet: Sheets: %s",str(sheet_names))
    obs_ws = wb.get_sheet_by_name(sheet)
    return wb,obs_ws

def create_metadata_sheet(metasheet,sheetname):
  """
  Create a sheet with metadata columns

  @param metasheet : empty worksheet
  @type  metasheet : openpyxl.Worksheet() instance

  @param sheetname : name to be given to sheet
  @type  sheetname : str

  @return: initialized metadata sheet
  """
  metasheet.title = sheetname
  metasheet.cell(column=0,row=0).value = "File"
  metasheet.cell(column=1,row=0).value = "Freq"
  metasheet.cell(column=2,row=0).value = "Pol"
  metasheet.cell(column=3,row=0).value = "BW"
  metasheet.cell(column=4,row=0).value = "IF mode"
  metasheet.cell(column=5,row=0).value = "First"
  metasheet.cell(column=6,row=0).value = "Last"
  metasheet.cell(column=7,row=0).value = "Start"
  metasheet.cell(column=8,row=0).value = "Stop"
  metasheet.cell(column=9,row=0).value = "Attn"
  return metasheet

def open_metadata_spreadsheet(filename,create_if_needed=True):
  """
  Load the observations metadata spreadsheet or create it.

  @param filename : filename of the spreadsheet

  @return: workbook instance
  """
  try:
    wb = openpyxl.load_workbook(filename)
  except IOError as details:
    logger.error("open_metadata_spreadsheet: loading failed with IO error: %s", str(details))
    raise IOError
  except AttributeError as details:
    logger.error("open_metadata_spreadsheet: loading failed with attribute error: %s", str(details))
    raise AttributeError
  except excel.InvalidFileException as details:
    logger.warning("open_metadata_spreadsheet: file does not exist.")
    if create_if_needed:
      logger.info("open_metadata_spreadsheet: creating workbook.")
      wb = openpyxl.Workbook()
      logger.debug("open_metadata_spreadsheet: sheets: %s", wb.get_sheet_names())
      obs_ws = wb.get_active_sheet()
      create_metadata_sheet(obs_ws,"Metadata")
    else:
      return None,None
  else:
    sheet_names = wb.get_sheet_names()
    logger.debug("open_metadata_spreadsheet: sheets: %s",str(sheet_names))
    obs_ws = wb.get_sheet_by_name('Metadata')
  return wb,obs_ws

def load_meta_sheet(wb,obs_ws):
  """
  Load the observations spreadsheet or create it.

  @param wb : session workbook
  @type  wb : openpyxl workbook instance

  @param obs_ws : metadata worksheet
  @param obs_ws : openpyxl worksheet instance

  @return: (worksheet column names,
            dictionary of column numbers)
  """
  obs_col_names = get_column_names(obs_ws)
  # make a reverse lookup table
  meta_column = {}
  logger.debug("load_meta_sheet: Worksheet %s columns: %s", obs_ws.title, str(obs_col_names))
  for col in list(obs_col_names.keys()):
    meta_column[obs_col_names[col]]  = get_column_id(obs_ws,obs_col_names[col])
  return obs_col_names, meta_column

def get_meta_data(meta_ws, meta_column, files):
  """
  Get metadata from session worksheet

  @param meta_ws : session worksheet
  @type  meta_ws : worksheet instance

  @param files : names of data files (which select rows)
  @type  files : list of str

  @return: (dictionary of freq indexed by filename,
            dictionary of polarization indexed by filename,
            datetime of start,
            mean datetime,
            datetime of end,
            doy,
            dictionary of solar_data)
  """
  freq = {}
  pol = {}
  # IFmode = {}
  first = {}
  last = {}
  start = {}
  stop = {}
  logger.debug("get_meta_data: from worksheet %s", meta_ws)
  logger.debug("get_meta_data: for files %s", files)
  for filename in files:
    bname = os.path.basename(filename)
    row = get_row_number(meta_ws, meta_column['File'], bname)
    logger.debug("get_meta_data: %s metadata are in row %d", bname, row)
    freq[bname]   = \
      meta_ws.cell(row=row, column=meta_column['Freq'] ).value
    pol[bname]    = \
      meta_ws.cell(row=row, column=meta_column['Pol']  ).value
    #IFmode[bname] = \
    #  meta_ws.cell(row=row, column=meta_column['IF mode']  ).value
    first[bname]  = \
      meta_ws.cell(row=row, column=meta_column['First']  ).value
    last[bname]   = \
      meta_ws.cell(row=row, column=meta_column['Last']  ).value
    start[bname]= \
      meta_ws.cell(row=row, column=meta_column['Start']).value
    stop[bname] = \
      meta_ws.cell(row=row, column=meta_column['Stop'] ).value
  mean_time = start[os.path.basename(files[0])] \
            + (stop[os.path.basename(files[0])]-start[os.path.basename(files[0])])/2
  doy = DT.day_of_year(mean_time.year, mean_time.month, mean_time.day)
  jd = DT.julian_date(mean_time.year,doy) + \
       (mean_time.hour + mean_time.minute/60.)/24.
  solar_data = solar.calc_solar(jd)
  # return freq,pol,IFmode,first,last,start,stop,mean_time,solar_data
  return freq,pol,first,last,start,stop,mean_time,solar_data

def get_file_freqs_and_pols(ws, meta_column, files):
  """
  Create a dictionary files indexed by frequency and polarization

  @param ws : worksheet with metadata

  @param files : list of filenames
  """
  #freq,pol,IFmode,first,last,start,stop,mean_time,solar_data = \
  #  get_meta_data(ws, meta_column, files)
  freq,pol,first,last,start,stop,mean_time,solar_data = \
    get_meta_data(ws, meta_column, files)
  filename_dict = {}
  for fn in files:
    filename = os.path.basename(fn)
    f = freq[filename]
    p = pol[filename]
    if f in filename_dict:
      filename_dict[f][p] = filename
    else:
      filename_dict[f] = {p: filename}

  fkeys = list(filename_dict.keys())
  fkeys.sort
  pkeys = ["LCP","RCP"]
  return fkeys, pkeys, filename_dict
